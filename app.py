# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
from typing import List, Dict, Tuple, Optional, Any
import joblib
import gc
import requests
import os
import zipfile
import xml.etree.ElementTree as ET
import html
from pathlib import Path

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias · API - Realizado por Johnathan Cortés",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"

CONCURRENT_REQUESTS          = 50
SIMILARITY_THRESHOLD_TONO    = 0.82
SIMILARITY_THRESHOLD_TITULOS = 0.93

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia han hay hubo habra
habria estoy esta estan estaba estaban estamos estan estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de","del","la","el","los","las","un","una","unos","unas","al","su","sus",
    "en","con","sin","por","para","sobre","ante","bajo","contra","desde",
    "entre","hacia","hasta","mediante","tras","y","o","u","e","lo","que","se",
    "como","donde","cuando","cual","cuyo","cuya","cuyos","cuyas",
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas","cada","todo","toda","todos","todas",
    "otro","otra","otros","otras","nuevo","nueva","nuevos","nuevas",
    "gran","grandes","mayor","mayores","menor","menores","mejor","mejores",
    "peor","peores","primer","primera","segundo","segunda","tercer","tercera",
    "más","mas","muy","tan","tanto","tanta","tantos","tantas",
    "mi","mis","tu","tus","nuestro","nuestra","nuestros","nuestras",
    "a","ha","he","ser","estar","haber","hacer","tener","poder","deber",
    "ir","dar","ver","saber","querer","llegar","pasar","decir","poner",
}

_TILDE_MAP = {
    "regulacion":"regulación","regulaciones":"regulaciones","innovacion":"innovación",
    "innovaciones":"innovaciones","tecnologia":"tecnología","tecnologias":"tecnologías",
    "tecnologica":"tecnológica","tecnologico":"tecnológico","educacion":"educación",
    "gestion":"gestión","administracion":"administración","informacion":"información",
    "comunicacion":"comunicación","comunicaciones":"comunicaciones","operacion":"operación",
    "operaciones":"operaciones","inversion":"inversión","inversiones":"inversiones",
    "expansion":"expansión","adquisicion":"adquisición","adquisiciones":"adquisiciones",
    "fusion":"fusión","fusiones":"fusiones","transicion":"transición",
    "transformacion":"transformación","digitalizacion":"digitalización",
    "automatizacion":"automatización","modernizacion":"modernización",
    "optimizacion":"optimización","implementacion":"implementación","evaluacion":"evaluación",
    "planificacion":"planificación","organizacion":"organización","atencion":"atención",
    "produccion":"producción","construccion":"construcción","distribucion":"distribución",
    "exportacion":"exportación","importacion":"importación","comercializacion":"comercialización",
    "negociacion":"negociación","negociaciones":"negociaciones","participacion":"participación",
    "colaboracion":"colaboración","asociacion":"asociación","integracion":"integración",
    "relacion":"relación","relaciones":"relaciones","situacion":"situación",
    "condicion":"condición","condiciones":"condiciones","solucion":"solución",
    "soluciones":"soluciones","prevencion":"prevención","proteccion":"protección",
    "fiscalizacion":"fiscalización","sancion":"sanción","sanciones":"sanciones",
    "investigacion":"investigación","investigaciones":"investigaciones","accion":"acción",
    "acciones":"acciones","direccion":"dirección","decision":"decisión",
    "decisiones":"decisiones","eleccion":"elección","elecciones":"elecciones",
    "votacion":"votación","aprobacion":"aprobación","legislacion":"legislación",
    "reclamacion":"reclamación","reclamaciones":"reclamaciones","obligacion":"obligación",
    "obligaciones":"obligaciones","inflacion":"inflación","tributacion":"tributación",
    "financiera":"financiera","financiero":"financiero","economica":"económica",
    "economico":"económico","economia":"economía","credito":"crédito",
    "creditos":"créditos","prestamo":"préstamo","prestamos":"préstamos",
    "interes":"interés","comision":"comisión","comisiones":"comisiones",
    "politica":"política","politicas":"políticas","politico":"político",
    "publica":"pública","publico":"público","estrategia":"estrategia",
    "estrategica":"estratégica","estrategico":"estratégico","logistica":"logística",
    "analisis":"análisis","diagnostico":"diagnóstico","indice":"índice",
    "vehiculo":"vehículo","vehiculos":"vehículos","electrico":"eléctrico",
    "electrica":"eléctrica","energia":"energía","energetica":"energética",
    "petroleo":"petróleo","mineria":"minería","agricola":"agrícola",
    "biologica":"biológica","ecologica":"ecológica","inclusion":"inclusión",
    "exclusion":"exclusión","pension":"pensión","pensiones":"pensiones",
    "jubilacion":"jubilación","compensacion":"compensación","remuneracion":"remuneración",
    "contratacion":"contratación","capacitacion":"capacitación","formacion":"formación",
    "certificacion":"certificación","habilitacion":"habilitación","autorizacion":"autorización",
    "concesion":"concesión","licitacion":"licitación","migracion":"migración",
    "poblacion":"población","recaudacion":"recaudación","asignacion":"asignación",
    "corporacion":"corporación","fundacion":"fundación","institucion":"institución",
    "instituciones":"instituciones","region":"región","unico":"único","unica":"única",
    "ultimo":"último","ultima":"última","proximo":"próximo","basico":"básico",
    "basica":"básica","historico":"histórico","historica":"histórica",
    "medico":"médico","medica":"médica","farmaceutica":"farmacéutica",
    "clinica":"clínica","numero":"número","telefono":"teléfono","telefonia":"telefonía",
    "movil":"móvil","moviles":"móviles","codigo":"código","informatica":"informática",
    "electronica":"electrónica","robotica":"robótica","ciberseguridad":"ciberseguridad",
    "trafico":"tráfico","transito":"tránsito","aereo":"aéreo","maritimo":"marítimo",
    "turistica":"turística","turistico":"turístico","gastronomia":"gastrónomía",
    "academica":"académica","academico":"académico","pedagogica":"pedagógica",
    "cientifica":"científica","cientifico":"científico","juridica":"jurídica",
    "juridico":"jurídico","constitucion":"constitución","resolucion":"resolución",
    "notificacion":"notificación","programacion":"programación","actualizacion":"actualización",
    "verificacion":"verificación","validacion":"validación","liquidacion":"liquidación",
    "facturacion":"facturación","evasion":"evasión","corrupcion":"corrupción",
    "deforestacion":"deforestación","contaminacion":"contaminación","conservacion":"conservación",
    "restauracion":"restauración","rehabilitacion":"rehabilitación","renovacion":"renovación",
    "ampliacion":"ampliación","inauguracion":"inauguración","celebracion":"celebración",
    "clasificacion":"clasificación","eliminacion":"eliminación","motivacion":"motivación",
    "satisfaccion":"satisfacción","reputacion":"reputación","disposicion":"disposición",
}

_ENIE_MAP = {
    "desempeno":"desempeño","desempenos":"desempeños","empeno":"empeño","empenos":"empeños",
    "ensenanza":"enseñanza","ensenanzas":"enseñanzas","diseno":"diseño","disenos":"diseños",
    "disenador":"diseñador","disenadora":"diseñadora","disenadores":"diseñadores",
    "nino":"niño","nina":"niña","ninos":"niños","ninas":"niñas","ninez":"niñez",
    "ano":"año","anos":"años","danio":"daño","danios":"daños","dano":"daño","danos":"daños",
    "danino":"dañino","danina":"dañina","montana":"montaña","montanas":"montañas",
    "espana":"España","espanol":"español","espanola":"española","espanoles":"españoles",
    "companero":"compañero","companera":"compañera","companeros":"compañeros","companeras":"compañeras",
    "compania":"compañía","companias":"compañías","acompanamiento":"acompanamiento",
    "cana":"caña","canas":"cañas","banio":"baño","banios":"baños","bano":"baño","banos":"baños",
    "pena":"peña","penas":"peñas","penon":"peñón","senor":"señor","senora":"señora",
    "senores":"señores","senoras":"señoras","senal":"señal","senales":"señales",
    "senalizacion":"señalización","pequeno":"pequeño","pequena":"pequeña",
    "pequenos":"pequeños","pequenas":"peñas","sueno":"sueño","suenos":"sueños",
    "dueno":"dueño","duena":"dueña","duenos":"dueños","duenas":"dueñas",
    "otono":"otoño","punio":"puño","punios":"puños","puno":"puño",
    "canon":"cañón","canones":"cañones","manana":"mañana","mananas":"mañanas",
    "cabana":"cabaña","cabanas":"cabañas","banera":"bañera","vinedo":"viñedo",
    "vinedos":"viñedos","rebano":"rebaño","rebanos":"rebaños","extrano":"extraño",
    "extrana":"extraña","extranos":"extraños","extranas":"extrañas",
    "enganio":"engaño","engano":"engaño","enganos":"engaños","tamanio":"tamaño",
    "tamano":"tamaño","tamanos":"tamaños","muneca":"muñeca","munecas":"muñecas",
    "cunado":"cuñado","cunada":"cuñada","cunados":"cuñados","albanil":"albañil",
    "albaniles":"albañiles","narino":"Nariño","quindio":"Quindío",
    "ibanez":"Ibáñez","nunez":"Núñez","munoz":"Muñoz","ordonez":"Ordóñez",
    "yanez":"Yáñez","castaneda":"Castañeda","penalosa":"Peñalosa",
    "vineta":"viñeta","vinetas":"viñetas","banado":"bañado","banada":"bañada",
    "rinon":"riñón","rinones":"riñones","panial":"pañal","paniales":"pañales",
    "panal":"pañal","panales":"pañales","arana":"araña","aranas":"arañas",
    "pestana":"pestaña","pestanas":"pestañas","guino":"guiño","guinos":"guiños",
    "munequera":"muñequera","lenador":"leñador","lenadores":"leñadores",
    "resena":"reseña","resenas":"reseñas","panuelo":"pañuelo","panuelos":"pañuelos",
    "companerismo":"compañerismo","desengano":"desengaño","lenio":"leño","leno":"leño",
}

def corregir_tildes(texto: str) -> str:
    if not texto: return texto
    palabras = texto.split()
    resultado = []
    for p in palabras:
        low = p.lower()
        if low in _TILDE_MAP:
            c = _TILDE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        elif low in _ENIE_MAP:
            c = _ENIE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        else:
            resultado.append(p)
    return " ".join(resultado)


# ======================================
# CSS
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
[data-testid="stTabs"] [data-testid="stTabsList"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;padding:4px!important;gap:4px!important;box-shadow:var(--shadow-sm)!important;margin-bottom:0.75rem!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]{font-family:'Google Sans',sans-serif!important;font-size:0.88rem!important;font-weight:500!important;color:var(--text2)!important;border-radius:var(--r)!important;padding:0.45rem 1.2rem!important;border:none!important;background:transparent!important;transition:var(--transition)!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]:hover{background:var(--s2)!important;color:var(--text)!important}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"]{background:var(--accent-bg)!important;color:var(--accent2)!important;border:1px solid var(--accent-bdr)!important;font-weight:700!important;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card.m-cost::before{background:linear-gradient(90deg,#f97316,#fb923c)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-icon.uz-region{background:#ecfdf5;color:#059669}
.upload-zone-icon.uz-internet{background:#eff6ff;color:#1a73e8}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
[data-testid="stTextInput"] input::placeholder,[data-testid="stTextArea"] textarea::placeholder{color:var(--text4)!important;font-size:0.85rem!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
[data-testid="stRadio"] label{font-family:'Google Sans Text',sans-serif!important;color:var(--text)!important;font-size:0.88rem!important;font-weight:400!important;}
[data-testid="stRadio"]{margin-bottom:0!important}
[data-testid="stRadio"]>div{gap:0!important}
[data-testid="stStatus"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;font-family:'Roboto Mono',monospace!important;font-size:0.8rem!important;}
[data-testid="stAlert"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;color:var(--text2)!important;font-size:0.85rem!important;padding:0.6rem 0.8rem!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 16px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
.cluster-info{background:var(--accent-bg);border:1px solid var(--accent-bdr);border-radius:var(--r);padding:0.5rem 0.8rem;margin:0.4rem 0;font-family:'Roboto Mono',monospace;font-size:0.68rem;color:var(--text2);line-height:1.6;}
.cluster-info b{color:var(--accent2);font-size:0.72rem}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
[data-testid="stSelectbox"]>div>div{font-family:'Google Sans Text',sans-serif!important;font-size:0.88rem!important;color:var(--text)!important;}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .upload-zone{grid-template-columns:1fr}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0

def get_embedding_cache():
    if '_emb_cache' not in st.session_state:
        st.session_state['_emb_cache'] = EmbeddingCache()
    return st.session_state['_emb_cache']


# ======================================
# Inicializador Seguro de Estado de Sesión
# ======================================
def init_session_state():
    """Garantiza la inicialización segura de variables antes de renderizar la app."""
    if 'tokens_input' not in st.session_state: 
        st.session_state['tokens_input'] = 0
    if 'tokens_output' not in st.session_state: 
        st.session_state['tokens_output'] = 0
    if 'tokens_embedding' not in st.session_state: 
        st.session_state['tokens_embedding'] = 0
    if '_emb_cache' not in st.session_state:
        st.session_state['_emb_cache'] = EmbeddingCache()


# ======================================
# Utilidades de Configuración Automática
# ======================================
def load_local_config():
    paths_to_try = [
        Path("Configuracion.xlsx"),
        Path("configuracion.xlsx"),
        Path("Config.xlsx"),
        Path("config.xlsx")
    ]
    for p in paths_to_try:
        if p.exists():
            return p
    base = Path(__file__).parent
    for f in base.iterdir():
        if f.suffix.lower() == '.xlsx' and 'config' in f.stem.lower():
            return f
    return None

def load_config(config_source):
    config_sheets = pd.read_excel(config_source, sheet_name=None, engine='openpyxl')
    region_map = pd.Series(
        config_sheets['Regiones'].iloc[:, 1].values,
        index=config_sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    internet_map = pd.Series(
        config_sheets['Internet'].iloc[:, 1].values,
        index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    return region_map, internet_map


# ======================================
# Función de Password robusta con Fallback local
# ======================================
def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Análisis La Cardio 🏥</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                try:
                    # Intenta recuperar desde los secretos del servidor
                    correct_pw = st.secrets.get("APP_PASSWORD", None)
                    if correct_pw is None:
                        correct_pw = "admin123"
                except Exception:
                    # Si st.secrets no está configurado (entorno local sin secrets.toml)
                    correct_pw = "admin123"
                
                if pw == correct_pw:
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            time.sleep(d)
            d *= 2

def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def string_norm_label(s):
    if not s: return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)

def clean_text(text):
    if not isinstance(text, str):
        return text
    return html.unescape(text).strip()

def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = html.unescape(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()

def normalize_title_for_comparison(title):
    """
    Normaliza el título de forma estricta, removiendo tildes usando unidecode y
    eliminando sufijos descriptivos o fuentes externas para agrupamientos precisos.
    """
    if not isinstance(title, str): 
        return ""
    # 1. Quitar tildes (ej: Séptima -> septima, séptima -> septima)
    cleaned = unidecode(title).strip()
    # 2. Remover fuentes externas finales (" | El Tiempo", " - Caracol", etc.)
    cleaned = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", cleaned).strip()
    # 3. Remover prefijos descriptivos de lugar/tag (ej: "Cundinamarca: ...")
    if ":" in cleaned:
        parts = cleaned.split(":", 1)
        suffix = parts[1].strip()
        if len(suffix) >= 10:
            cleaned = suffix
    # 4. Pasar a minúsculas y omitir caracteres no alfanuméricos
    return re.sub(r"\W+", " ", cleaned).lower().strip()

def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
        'television': 'Televisión', 'televisión': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }.get(t, str(tipo_raw).strip().title() or "Otro")

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    if 'e' in s.lower():
        s = s.replace(',', '.')
    else:
        if ',' in s and '.' in s:
            if s.rfind('.') < s.rfind(','):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            parts = s.split(',')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0,')):
                s = s.replace(',', '')
            else:
                s = s.replace(',', '.')
        elif '.' in s:
            parts = s.split('.')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0.')):
                s = s.replace('.', '')
    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None

def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {t}. {t}. {r}"[:max_len]

# ======================================
# Limpieza de Menciones Específica
# ======================================
def limpiar_mencion(mencion_str):
    """
    Limpia y estandariza el nombre de la marca en Menciones - Empresa.
    Quita el prefijo 'La Cardio 26 - ' si se encuentra en la mención.
    """
    if not isinstance(mencion_str, str):
        return mencion_str
    # Quitar cualquier ocurrencia de "La Cardio 26 - " de forma case-insensitive
    cleaned = re.sub(r'(?i)\bLa\s+Cardio\s+26\s*-\s*', '', mencion_str)
    return cleaned.strip()

# ======================================
# Filtro de Titulares de Última Hora / Genéricos y No Salud
# ======================================
def es_titular_generico(titulo: str) -> bool:
    """
    Verifica localmente si un titular contiene cadenas comunes de resúmenes diarios,
    titulares sin contexto o de última hora para clasificarlos directamente.
    """
    if not titulo: return False
    t = unidecode(str(titulo).lower())
    patrones = [
        r"titulares\s+de\s+hoy",
        r"titulares\s+del\s+dia",
        r"titulares\s+ultima\s+hora",
        r"ultima\s+hora",
        r"resumen\s+de\s+noticias",
        r"las\s+noticias\s+de\s+hoy",
        r"noticias\s+del\s+dia",
        r"lo\s+que\s+debe\s+saber",
        r"boletin\s+de\s+noticias",
        r"titulares\s+de\s+la\s+manana",
        r"titulares\s+de\s+la\s+tarde",
        r"titulares\s+de\s+la\s+noche",
        r"titulares\s+de\s+prensa",
        r"resumen\s+informativo",
        r"titulares\s+de\s+noticias",
        r"titulares\s+de\s+ultima\s+hora",
        r"noticias\s+en\s+60\s+segundos"
    ]
    for p in patrones:
        if re.search(p, t):
            return True
    return False

def es_noticia_no_salud(titulo: str) -> bool:
    """
    Identifica de forma local noticias que traten exclusivamente de accidentes, obras, cierres
    viales, nacimientos, abogados o temas sin relación con salud médica / reputación hospitalaria.
    """
    if not titulo: return False
    t = unidecode(str(titulo).lower())
    
    # Palabras clave de eventos ajenos a salud en el título
    palabras_clave = [
        "cierre", "obras", "corredor verde", "septima", "séptima", "calle", "via", "vía", 
        "accidente", "choque", "volcamiento", "nacimiento", "abogado", "transito", "tránsito",
        "movilidad", "desvío", "desvio", "pavimentacion", "pavimentación", "embotellamiento"
    ]
    
    if any(kw in t for kw in palabras_clave):
        # Aseguramos que no existan excepciones médicas en el título que sí requieran análisis de salud
        excepciones_salud = [
            "cirugia", "cirugía", "medico", "médico", "doctor", "tratamiento", "clinica", "clínica", 
            "hospital", "salud", "cardio", "trasplante", "paciente", "oncologia", "oncología", 
            "neurologia", "neurología", "acreditacion", "acreditación"
        ]
        if not any(ex in t for ex in excepciones_salud):
            return True
    return False

def es_suceso_excluido_politica(titulo: str) -> bool:
    """
    Verifica si el título contiene temas de la coyuntura política y personajes específicos
    (expresidentes, Gaviria, Miguel Uribe, atentados políticos, Wendy Sepúlveda o boletines genéricos)
    para clasificarlos directamente como Neutro, Sucesos y Otras de forma local e infalible.
    """
    if not titulo: return False
    t = unidecode(str(titulo).lower())
    
    # Comprobar el atentado electoral específico
    if "atentado que cambio historia de los comicios" in t or "comicios a la presidencia" in t:
        return True
        
    # Palabras clave y personajes para forzar a Sucesos / Otras
    patrones = [
        "cesar gaviria", "gaviria", "expresidente", "Noticias en 60 Segundos","Yulixa Toloza","Velatón en Cali", "Mujer murió ",
        "miguel uribe", "miguel uribe turbay", "Noticias en", "Un año desde la tragedia", "Una enfermera murió ", "BALEARON A ",
        "caviria", "cavi ria", "caviria sigue hospitalizado", "cavi ria sigue hospitalizado","Tragedia en","ataque armado",
        "wendy sepulveda", "noticias del", "noticias nacionales", "noticias en ","Velatón en", "Noticias en ", "FUE ARROLLADA",
        "atentado contra miguel uribe","decencia ha entrado en pausa"
    ]
    
    for p in patrones:
        p_norm = unidecode(p.lower())
        if p_norm in t:
            return True
            
    return False


# ======================================
# Estructuras de Datos Avanzadas
# ======================================
class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n): c[self.find(i)].append(i)
        return dict(c)

def agrupar_textos_similares(textos, umbral):
    if not textos: return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2: return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g = defaultdict(list)
    for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))

def agrupar_por_titulo_similar(titulos):
    """
    Agrupa títulos homólogos basándose en un algoritmo que combina correspondencia
    estricta unidecode y similitud Jaccard de tokens.
    """
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    
    def son_similares(s1, s2):
        if not s1 or not s2: return False
        if s1 == s2: return True
        # Ratio de SequenceMatcher
        seq_ratio = SequenceMatcher(None, s1, s2).ratio()
        if seq_ratio >= 0.85: return True
        
        # Jaccard de palabras
        w1, w2 = set(s1.split()), set(s2.split())
        if w1 and w2:
            jaccard = len(w1 & w2) / len(w1 | w2)
            if jaccard >= 0.80: return True
        return False

    for i in range(len(norm)):
        if i in used or not norm[i]: continue
        grp = [i]
        used.add(i)
        for j in range(i + 1, len(norm)):
            if j in used or not norm[j]: continue
            if son_similares(norm[i], norm[j]):
                grp.append(j)
                used.add(j)
        if len(grp) >= 1:
            grupos[gid] = grp
            gid += 1
    return grupos

def seleccionar_representante(indices, textos):
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos: return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]

def get_embeddings_batch(textos, batch_size=100):
    if not textos: return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing: return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_embedding'] += (u.get('total_tokens') if isinstance(u, dict) else getattr(u, 'total_tokens', 0)) or 0
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                emb = d["embedding"]
                resultados[oi] = emb
                cache.put(textos[oi], emb)
        except:
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    emb = r["data"][0]["embedding"]
                    resultados[oi] = emb
                    cache.put(textos[oi], emb)
                except:
                    pass
    return resultados


# ======================================
# ANALIZADOR INTELIGENTE UNIFICADO - MULTI-HILOS ESTABLE (ThreadPoolExecutor)
# ======================================
class ClasificadorNoticiasInteligente:
    def __init__(self, marca_principal, aliases):
        self.marca_principal = marca_principal.strip()
        self.aliases = [a.strip() for a in (aliases or []) if a.strip()]
        self._all_names = [self.marca_principal.lower()] + [a.lower() for a in self.aliases]

    def _menciona_marca(self, texto, marca_especifica):
        """
        Detecta la presencia de marcas con lógica de sinónimos inteligente
        (ej: FCV <=> Cardiovascular, La Cardio <=> CardioInfantil, Country <=> Clínica del Country).
        """
        t = unidecode(texto.lower())
        m = unidecode(str(marca_especifica).lower().strip()) if marca_especifica else ""
        if not m or m in ("nan", "n/a", "none", "-"):
            return any(n in t for n in self._all_names)
            
        # Diccionario de grupos de sinónimos de marcas comunes para homologar la detección de menciones
        sinonimos_marcas = {
            "cardiovascular": ["cardiovascular", "fcv", "fundacion cardiovascular", "clinica cardiovascular"],
            "fcv": ["cardiovascular", "fcv", "fundacion cardiovascular", "clinica cardiovascular"],
            "country": ["country", "clinica del country", "clínica del country"],
            "la cardio": ["la cardio", "lacardio", "cardioinfantil", "cardio infantil", "fundacion cardioinfantil"],
            "cardioinfantil": ["la cardio", "lacardio", "cardioinfantil", "cardio infantil", "fundacion cardioinfantil"],
            "cardio infantil": ["la cardio", "lacardio", "cardioinfantil", "cardio infantil", "fundacion cardioinfantil"]
        }
        
        # Evaluar coincidencia de sinónimos si la marca corresponde a uno de los grupos definidos
        for canon, grupo in sinonimos_marcas.items():
            if canon in m:
                if any(syn in t for syn in grupo):
                    return True
                    
        return m in t

    def _es_relevancia_positiva_salud(self, titulo, resumen) -> bool:
        """
        Detecta localmente palabras clave del sector salud que denotan alta excelencia,
        logros, cirugías exitosas, avances o premios para guiar el tono hacia 'Positivo'.
        """
        t_r = unidecode((str(titulo) + " " + str(resumen)).lower())
        indicadores = [
            "acreditacion", "acreditación", "certificacion", "certificación", "reconocimiento",
            "reconocida", "reconocido", "premio", "galardon", "galardón", "hilo", "hito",
            "innovacion", "innovación", "innovadoras", "innovadora", "innovador", "innovadores",
            "avance", "avances", "tecnologia", "tecnología", "cirugia", "cirugía", "operacion", "operación",
            "procedimiento exitoso", "pionero", "pionera", "mejor clinica", "mejor clínica",
            "ranking", "escalafon", "escalafón", "exito", "éxito", "trasplante", "salva vidas",
            "primer puesto", "lider", "líder", "destaca", "destacan"
        ]
        return any(ind in t_r for ind in indicadores)

    def _analizar_llm_sync(self, texto, marca_especifica):
        """Llamada síncrona a la API de OpenAI, ideal para ejecución segura en ThreadPoolExecutor."""
        marca_target = str(marca_especifica).strip() if (marca_especifica and str(marca_especifica).strip() not in ("", "nan", "N/A", "-")) else self.marca_principal
        
        # Si el texto de la noticia no contiene mención de la marca evaluada, devolvemos valores por defecto neutros
        if not self._menciona_marca(texto, marca_target):
            return {
                "tono": "Neutro",
                "categoria": "Sector",
                "narrativa": "Otras"
            }

        prompt = (
            f"Eres un experto analista de reputación, prensa y posicionamiento de marcas en el sector salud de Colombia.\n"
            f"Tu tarea consiste en realizar un análisis de prensa inteligente, contextual y de alta precisión para la marca '{marca_target}' en la siguiente noticia.\n\n"
            f"TEXTO DE LA NOTICIA a evaluar:\n{texto[:1800]}\n\n"
            f"ENTIDAD BAJO ANÁLISIS EN ESTE CASO:\n'{marca_target}'\n\n"
            f"--- INSTRUCCIONES DE ENFOQUE CRÍTICO (FUNDAMENTAL) ---\n"
            f"Tu análisis debe centrarse de manera estricta en la marca '{marca_target}' y en CÓMO se le asocia en la noticia.\n"
            f"Por ejemplo, si la noticia aborda un logro de innovación médica general de un competidor, pero la entidad '{marca_target}' NO es la que realiza el avance, la narrativa de esta fila NO debe ser 'Innovación + Desarrollo' sino 'Otras'. Sé analítico, preciso y contextual.\n\n"
            f"1. TONO DE REPUTACIÓN (En relación directa con '{marca_target}'):\n"
            f"Evalúa cómo afecta el artículo a la imagen de '{marca_target}':\n"
            f"- Positivo: Reconocimientos, premios, acreditaciones de calidad (nacionales o internacionales), cirugías exitosas, innovaciones médicas, aportes científicos, expansión o hitos que exalten y beneficien la imagen corporativa de '{marca_target}' en salud.\n"
            f"- Negativo: Demandas penales o civiles, fallas clínicas u operativas, quejas médicas, crisis institucionales, investigaciones o reclamaciones directas contra '{marca_target}'.\n"
            f"- Neutro: Menciones secundarias, informativas de contexto sectorial o de pasillo sin connotación reputacional directa de éxito o fracaso.\n\n"
            f"2. CATEGORÍA (Involucramiento institucional - Escribe exactamente uno de estos nombres):\n"
            f"- Sucesos: Acciones que ocurren en donde aparece mi marca, pero no se encuentran ligados a mi objetivo de negocio. Ej: Heridos en la balacera en Usaquén, cierres viales u obras viales donde se asocia la marca por cercanía.\n"
            f"- Core: Servicios direccionados a mi estrategia de negocio cardiovascular y trasplantes (exclusivo para cuando involucra cardiovascular/trasplante en '{marca_target}'). Ej: Síntomas de un infarto.\n"
            f"- Especialidades: Todas las otras especialidades médicas que se ofrecen en LaCardio (neurología, pediatría, oncología, neurología, etc.). Ej: Enfermedades neurológicas.\n"
            f"- Ranking: Menciones en las diferentes rankings, mediciones de reputación o escalafones. Ej: Top de las marcas colombianas P&M.\n"
            f"- Sector: Las menciones que se relacionan al sector salud en general, embargos, problemas de EPS o aspectos noticiosos que LaCardio debe tener en cuenta. Ej: Crisis en el sector salud.\n"
            f"- Reforma: Las menciones que van relacionadas a los lineamientos normativos que ocurren en el entorno y por los cuales la marca puede verse afectada (reforma a la salud, leyes, etc.). Ej: Activan la reforma de salud.\n"
            f"- Corporativo: Acciones que ocurren de la marca en donde corresponde a participaciones, reconocimientos o menciones como resultados de una alianza, asambleas o asocio corporativo. Ej: Encuentro Latidos Futuros 4.\n\n"
            f"3. NARRATIVAS (Rol del mensaje estratégico - Escribe exactamente uno de estos nombres):\n"
            f"- Sostenibilidad: Contenido relacionado a las acciones de sostenibilidad, voluntariado y propósito social. Ej: Donaciones de libros - Filbo 2026.\n"
            f"- Excelencia médica: Contenido relacionado con la experticia médica en las diferentes áreas de servicio, acreditaciones y calidad. Ej: Reacreditación de la Joint Commission International.\n"
            f"- Innovación + Desarrollo: Contenido relacionado con novedades, investigación, apertura de servicios, innovación en equipos y tratamientos realizados por '{marca_target}'. Ej: Producción científica Nature Index 2025.\n"
            f"- Marca empleadora: Contenido relacionado a las acciones, bienestar, perfiles y logros de nuestros colaboradores. Ej: Jaime Fernandez, reconocido médico es orgullo cardio.\n"
            f"- Portafolio: Contenido relacionado con los diferentes servicios médicos generales o consejos de hábitos saludables. Ej: Beneficios de la actividad física.\n"
            f"- Otras: Contenido relacionado con los diferentes servicios médicos pero de carácter puramente referencial, marketing sensorial o menciones secundarias. Ej: Referencial, Marketing sensorial.\n\n"
            f"Genera estrictamente un objeto JSON plano sin introducciones ni marcas de formato secundarias, exactamente de esta forma:\n"
            f'{{"tono": "Positivo|Negativo|Neutro", '
            f'"categoria": "Sucesos|Core|Especialidades|Ranking|Sector|Reforma|Corporativo", '
            f'"narrativa": "Sostenibilidad|Excelencia médica|Innovación + Desarrollo|Marca empleadora|Portafolio|Otras"}}'
        )

        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            
            resultado = json.loads(resp.choices[0].message.content)
            tono = str(resultado.get("tono", "Neutro")).strip().title()
            cat = str(resultado.get("categoria", "Sector")).strip().title()
            nar = str(resultado.get("narrativa", "Otras")).strip()
            
            valid_tonos = {"Positivo", "Negativo", "Neutro"}
            valid_cats = {"Sucesos", "Core", "Especialidades", "Ranking", "Sector", "Reforma", "Corporativo"}
            valid_nars = {"Sostenibilidad", "Excelencia médica", "Innovación + Desarrollo", "Marca empleadora", "Portafolio", "Otras"}
            
            if tono not in valid_tonos: tono = "Neutro"
            if cat not in valid_cats: cat = "Sector"
            if nar not in valid_nars:
                if "innovacion" in nar.lower() or "desarrollo" in nar.lower(): nar = "Innovación + Desarrollo"
                elif "excelencia" in nar.lower(): nar = "Excelencia médica"
                elif "marca" in nar.lower() or "empleador" in nar.lower(): nar = "Marca empleadora"
                else: nar = "Otras"
                
            return {"tono": tono, "categoria": cat, "narrativa": nar}
        except Exception:
            return {"tono": "Neutro", "categoria": "Sector", "narrativa": "Otras"}

    def procesar_lote(self, textos, pbar, resumenes, titulos, menciones):
        """Procesa el lote usando ThreadPoolExecutor nativo, asegurando estabilidad total en Streamlit."""
        n = len(textos)
        txts = textos.tolist()
        
        pbar.progress(0.05, "Agrupando estrictamente por títulos para consistencia absoluta de metadatos...")
        
        # 1. Agrupamiento exclusivo basado en similitud estricta del Título (unidecode + lower + Jaccard de palabras)
        grupos_titulos = agrupar_por_titulo_similar(titulos.tolist())
        
        # 2. Agrupar por combinación (GrupoTítuloId, MencionNormalizada)
        # Esto asegura de forma absoluta que noticias similares con la misma marca hereden idéntico análisis de metadatos
        grupos_por_mencion = defaultdict(list)
        for gid, idxs in grupos_titulos.items():
            for idx in idxs:
                menc = str(menciones.iloc[idx]).strip()
                menc_norm = norm_key(menc)
                grupos_por_mencion[(gid, menc_norm)].append(idx)
                
        llaves_subgrupo = list(grupos_por_mencion.keys())
        
        reps_mencion = []
        for k in llaves_subgrupo:
            idxs = grupos_por_mencion[k]
            rep_idx, rep_txt = seleccionar_representante(idxs, txts)
            menc_especifica = str(menciones.iloc[rep_idx]).strip()
            rep_titulo = str(titulos.iloc[rep_idx]).strip()
            reps_mencion.append((rep_txt, menc_especifica, rep_titulo))
            
        rpg = {}
        subgrupos_a_evaluar = []
        
        # Validación de Titular Genérico / No-Salud / Política Excluidos de forma local antes de invocar la API
        for idx, (txt, menc, r_title) in enumerate(reps_mencion):
            key = llaves_subgrupo[idx]
            if es_titular_generico(r_title):
                rpg[key] = {
                    "tono": "Neutro",
                    "categoria": "Sector",
                    "narrativa": "Otras"
                }
            elif es_noticia_no_salud(r_title) or es_suceso_excluido_politica(r_title):
                rpg[key] = {
                    "tono": "Neutro",
                    "categoria": "Sucesos",
                    "narrativa": "Otras"
                }
            else:
                subgrupos_a_evaluar.append((key, txt, menc))
                
        # Llamadas en paralelo síncronas usando ThreadPoolExecutor (Seguridad ante bucles de Streamlit)
        if subgrupos_a_evaluar:
            with ThreadPoolExecutor(max_workers=15) as executor:
                futures_to_key = {
                    executor.submit(self._analizar_llm_sync, txt, menc): key 
                    for key, txt, menc in subgrupos_a_evaluar
                }
                
                total_tareas = len(futures_to_key)
                for i, future in enumerate(as_completed(futures_to_key)):
                    key = futures_to_key[future]
                    try:
                        rpg[key] = future.result()
                    except Exception:
                        rpg[key] = {"tono": "Neutro", "categoria": "Sector", "narrativa": "Otras"}
                    pbar.progress(0.1 + 0.85 * (i + 1) / total_tareas, f"Analizando noticias {i + 1}/{total_tareas}")
        
        final = [None] * n
        for k, idxs in grupos_por_mencion.items():
            r = rpg.get(k, {"tono": "Neutro", "categoria": "Sector", "narrativa": "Otras"})
            
            # Obtener metadatos base evaluados del representante del subgrupo
            rep_idx = idxs[0]
            t_rep = titulos.iloc[rep_idx]
            r_rep = resumenes.iloc[rep_idx]
            
            # Aplicar reglas locales heurísticas una única vez sobre el representante del grupo
            rule_cat, rule_nar = self._clasificar_con_reglas_locales(t_rep, r_rep)
            cat_final = r.get("categoria")
            nar_final = r.get("narrativa")
            tono_final = r.get("tono")
            
            # Forzar tono POSITIVO para hitos médicos, reconocimientos, cirugías o innovaciones clave en el sector salud
            if self._es_relevancia_positiva_salud(t_rep, r_rep) and tono_final == "Neutro":
                # Validar que no sea una noticia excluida de tránsito o política antes de potenciar el tono
                if not es_noticia_no_salud(t_rep) and not es_suceso_excluido_politica(t_rep) and not es_titular_generico(t_rep):
                    tono_final = "Positivo"
            
            if cat_final == "Sector" and rule_cat in ("Core", "Especialidades", "Ranking", "Reforma", "Sucesos", "Corporativo"):
                cat_final = rule_cat
            if nar_final == "Otras" and rule_nar and rule_nar != "Otras":
                nar_final = rule_nar
                
            # Sobrescribir de manera estricta si es clasificado localmente como noticia de tránsito, política o boletín genérico
            if es_titular_generico(t_rep):
                tono_final = "Neutro"
                cat_final = "Sector"
                nar_final = "Otras"
            elif es_noticia_no_salud(t_rep) or es_suceso_excluido_politica(t_rep):
                tono_final = "Neutro"
                cat_final = "Sucesos"
                nar_final = "Otras"
                
            # Asignar de manera estrictamente idéntica a todos los miembros del subgrupo (Garantiza consistencia absoluta)
            for i in idxs:
                final[i] = {
                    "tono": tono_final,
                    "categoria": cat_final,
                    "narrativa": nar_final
                }
                
        pbar.progress(1.0, "Análisis de noticias finalizado")
        return final

    def _clasificar_con_reglas_locales(self, titulo, resumen) -> Tuple[Optional[str], Optional[str]]:
        t_r = (str(titulo) + " " + str(resumen)).lower()
        
        core_kw = ["infarto", "trasplante", "corazón", "corazon", "cardio", "hemodinamia", "marcapasos", "válvula", "valvula", "arritmia", "miocardio", "arteria", "cardiología", "cardiologia", "cardiovascular"]
        esp_kw = ["neurología", "neurologia", "ortopedia", "pediatría", "pediatria", "ginecología", "ginecologia", "anestesiología", "anestesiologia", "nefrología", "nefrologia", "dermatología", "dermatologia", "urología", "urologia", "oftalmología", "oftalmologia", "odontología", "consulta externa", "urgencias", "oncología", "oncologia"]
        rank_kw = ["ranking", "escalafón", "escalafon", "merco", "américa economía", "america economia", "p&m", "marcas colombianas", "prestigio", "medición", "monitor", "las mas innovadoras", "las más innovadoras"]
        ref_kw = ["reforma a la salud", "reforma de salud", "proyecto de ley", "senado", "debate de la reforma", "ley de salud"]
        sect_kw = ["crisis en el sector", "crisis de la salud", "eps", "minsalud", "adres", "embargo", "superintendencia de salud", "supersalud", "red hospitalaria", "clínicas", "escasez de medicamentos"]
        corp_kw = ["latidos futuros", "alianza", "convenio", "acreditación", "acreditacion", "reconocimiento corporativo", "junta directiva", "asamblea de socios", "junta", "recibe maxima", "recibe máxima"]
        suc_kw = ["balacera", "robo", "atraco", "herido", "choque", "accidente", "incendio", "capturado", "policía", "policia", "homicidio"]
        
        matched_cat = None
        if any(k in t_r for k in core_kw): matched_cat = "Core"
        elif any(k in t_r for k in esp_kw): matched_cat = "Especialidades"
        elif any(k in t_r for k in rank_kw): matched_cat = "Ranking"
        elif any(k in t_r for k in ref_kw): matched_cat = "Reforma"
        elif any(k in t_r for k in sect_kw): matched_cat = "Sector"
        elif any(k in t_r for k in corp_kw): matched_cat = "Corporativo"
        elif any(k in t_r for k in suc_kw): matched_cat = "Sucesos"
        
        sost_kw = ["sostenibilidad", "propósito social", "proposito social", "donación", "donacion", "filbo", "brigada", "responsabilidad social"]
        exc_kw = ["excelencia", "reacreditación", "reacreditacion", "joint commission", "jci", "experticia", "calidad médica", "alta complejidad", "maxima acreditacion", "máxima acreditación"]
        inn_kw = ["innovación", "innovacion", "desarrollo", "nature index", "investigación", "investigacion", "tecnología", "tecnologia", "patente", "telemedicina", "da vinci", "robot", "innovadoras"]
        emp_kw = ["colaborador", "empleado", "orgullo cardio", "talento humano", "bienestar", "enfermera", "médico es orgullo", "medico es orgullo"]
        port_kw = ["actividad física", "actividad fisica", "chequeo", "consejos de salud", "vacunación", "vacunacion", "nutrición", "nutricion"]
        
        matched_nar = None
        if any(k in t_r for k in sost_kw): matched_nar = "Sostenibilidad"
        elif any(k in t_r for k in exc_kw): matched_nar = "Excelencia médica"
        elif any(k in t_r for k in inn_kw): matched_nar = "Innovación + Desarrollo"
        elif any(k in t_r for k in emp_kw): matched_nar = "Marca empleadora"
        elif any(k in t_r for k in port_kw): matched_nar = "Portafolio"
        else: matched_nar = "Otras"
        
        return matched_cat, matched_nar


# ======================================
# Duplicados y Excel
# ======================================
def _normalizar_url(url: str) -> str:
    """Normaliza URLs eliminando el protocolo http/https y www para una comparación fiable."""
    if not url: return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url

def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue

        tipo    = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        medio   = norm_key(row.get(km["medio"], ""))

        streaming_url_raw = row.get(km["link_streaming"])
        if isinstance(streaming_url_raw, dict):
            streaming_url_raw = streaming_url_raw.get("url")
            
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(str(streaming_url_raw))
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_streaming[sk]].get(km["idnoticia"], "")
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li = row.get(km["link_nota"])
            url = li.get("url") if isinstance(li, dict) else li
            if url and mencion:
                url_norm = _normalizar_url(str(url))
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km["hora"], "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta  = normalize_title_for_comparison(processed[a].get(km["titulo"]))
                tb_ = normalize_title_for_comparison(processed[b].get(km["titulo"]))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a][km["idduplicada"]]  = processed[b].get(km["idnoticia"], "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b][km["idduplicada"]]  = processed[a].get(km["idnoticia"], "")

    return processed

def read_and_normalize_dossier(sheet, region_map, internet_map):
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(row):
                cell = row[i]
                val = cell.value
                url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                if url:
                    row_data[h] = {"value": val or "Link", "url": url}
                else:
                    row_data[h] = val
        rows.append(row_data)

    df = pd.DataFrame(rows)

    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    
    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(tipo_medio_map)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    if 'Medio' in df.columns:
        raw_medios_clean = df['Medio'].astype(str).str.lower().str.strip()
        df['Región'] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia'] = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha'] = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora'] = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor'] = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina'] = df.get('Nro. Pagina', pd.Series(dtype=str))
    
    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión'] = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión'] = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av = df.get('CPE', pd.Series([np.nan] * len(df)))
    cpe_grafica = df.get('Valor de Nota', pd.Series([np.nan] * len(df)))
    df['CPE'] = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier'] = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono'] = df.get('Tono', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    df['Categoría'] = df.get('Categoría', df.get('Categoria', df.get('Tematica', df.get('Tema', pd.Series(dtype=str))))).astype(str).apply(clean_text)
    df['Narrativas'] = df.get('Narrativas', df.get('Narrativa', df.get('Subtema', pd.Series(dtype=str)))).astype(str).apply(clean_text)

    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    url_nota_av = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df))))
    url_streaming = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df)))
    
    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})
                
    df['Link Nota'] = link_nota_final

    url_nota_raw = df.get('URL Nota', pd.Series([''] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)
            
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    menciones_grafica = df.get('Empresa rel.', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    # Aplicar la limpieza del prefijo "La Cardio 26 - "
    df['Menciones - Empresa'] = df['Menciones - Empresa'].apply(limpiar_mencion)

    return df

def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Tier", "Audiencia", "Tono", "Tono IA", "Categoría", "Narrativas",
        "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada"
    ]
    NUM = {"ID Noticia", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia"}
    ws.append(ORDER)
    
    font_hyperlink = Font(color="0563C1", underline="single")
    align_left = Alignment(horizontal='left')
    font_header = Font(bold=True)
    
    for i, col_name in enumerate(ORDER, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = font_header

    col_idx_map = {name: ORDER.index(name) + 1 for name in ORDER}
        
    for row in rows:
        tk = km.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        
        out, links = [], {}
        for ci, h in enumerate(ORDER, start=1):
            dk = km.get(norm_key(h), norm_key(h))
            val = row.get(h)
            cv = None
            
            if h == 'Fecha' and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val) if val is not None else None
            elif h in NUM:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    links[ci] = val
                else:
                    cv = str(val)
            out.append(cv)
        ws.append(out)
        
        current_row = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current_row, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
            
        date_col_idx = ORDER.index("Fecha") + 1
        date_cell = ws.cell(row=current_row, column=date_col_idx)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = 'DD/MM/YYYY'
            
        cols_millares = ["Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "Tier", "Audiencia"]
        for col_name in cols_millares:
            col_idx = col_idx_map[col_name]
            cell = ws.cell(row=current_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

        cpe_idx = col_idx_map["CPE"]
        cpe_cell = ws.cell(row=current_row, column=cpe_idx)
        if isinstance(cpe_cell.value, (int, float)):
            cpe_cell.number_format = '$#,##0'
            
    for i, col_name in enumerate(ORDER, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ['Título', 'Resumen - Aclaracion']:
            ws.column_dimensions[letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================================
# Proceso principal (Análisis Completo - Ejecución síncrona sin asyncio.run)
# ======================================
def run_full_process(df_file, bn, ba, tpkl, epkl, mode, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    t0 = time.time()
    
    if "API" in mode:
        try:
            openai.api_key=st.secrets["OPENAI_API_KEY"]
        except:
            st.error("OPENAI_API_KEY no encontrado en st.secrets.")
            st.stop()
            
    with st.status("Paso 1 · Carga de Configuración y Dossier", expanded=True) as s:
        config_path = load_local_config()
        if not config_path:
            st.error("❌ No se encontró el archivo 'Configuracion.xlsx' en el repositorio. Asegúrate de incluirlo en la raíz.")
            st.stop()
            
        region_map, internet_map = load_config(config_path)
        
        wb_in = load_workbook(df_file, data_only=True)
        df_normalized = read_and_normalize_dossier(wb_in.active, region_map, internet_map)
        
        # Expansión por ; en Menciones - Empresa
        rows_expanded = []
        for idx, row_series in df_normalized.iterrows():
            menciones = [m.strip() for m in str(row_series['Menciones - Empresa']).split(';') if m.strip()]
            if not menciones:
                row_dict = row_series.to_dict()
                row_dict['Menciones - Empresa'] = ""
                row_dict['original_index'] = idx
                row_dict['is_duplicate'] = False
                rows_expanded.append(row_dict)
            else:
                for m in menciones:
                    row_dict = row_series.to_dict()
                    row_dict['Menciones - Empresa'] = m
                    row_dict['original_index'] = idx
                    row_dict['is_duplicate'] = False
                    rows_expanded.append(row_dict)

        km = {
            "idnoticia": "ID Noticia",
            "fecha": "Fecha",
            "hora": "Hora",
            "medio": "Medio",
            "tipodemedio": "Tipo de Medio",
            "seccion_programa": "Sección - Programa",
            "region": "Región",
            "titulo": "Título",
            "autor_conductor": "Autor - Conductor",
            "nro_pagina": "Nro. Pagina",
            "dimension": "Dimensión",
            "duracion_caracteres": "Duración - Nro. Caracteres",
            "cpe": "CPE",
            "tier": "Tier",
            "audiencia": "Audiencia",
            "tono": "Tono",
            "tonoiai": "Tono IA",
            "tema": "Categoría",
            "categoria": "Categoría",
            "subtema": "Narrativas",
            "narrativas": "Narrativas",
            "link_nota": "Link Nota",
            "resumen": "Resumen - Aclaracion",
            "link_streaming": "Link (Streaming - Imagen)",
            "menciones": "Menciones - Empresa",
            "idduplicada": "ID duplicada"
        }
        
        rows = detectar_duplicados_avanzado(rows_expanded, km)
        for row in rows:
            if row["is_duplicate"]:
                row["Tono IA"] = "Duplicada"
                row[km["tema"]] = "-"
                row[km["subtema"]] = "-"
                
        s.update(label="✓ Paso 1 completado", state="complete")
        
    with st.status("Paso 2 · Normalización", expanded=True) as s:
        s.update(label="✓ Paso 2 · Mapeos y normalizaciones aplicados", state="complete")
        
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    
    if ta:
        df = pd.DataFrame(ta)
        df["_txt"] = df.apply(
            lambda r: texto_para_embedding(str(r.get(km["titulo"], "")), str(r.get(km["resumen"], ""))),
            axis=1
        )
        with st.status("Embeddings...", expanded=True) as s:
            _ = get_embeddings_batch(df["_txt"].tolist())
            s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
            
        with st.status("Paso 3 · Análisis Contextual Integrado (Tono, Categoría y Narrativas)", expanded=True) as s:
            pb = st.progress(0)
            if "API" in mode:
                # El análisis inteligente unificado evalúa tono, categoría y narrativa
                # de forma integrada por cada combinación de contenido + mención de marca única
                resultados_analisis = ClasificadorNoticiasInteligente(bn, ba).procesar_lote(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]], df[km["menciones"]]
                )
                df[km["tonoiai"]] = [r["tono"] for r in resultados_analisis]
                df[km["tema"]] = [r["categoria"] for r in resultados_analisis]
                df[km["subtema"]] = [r["narrativa"] for r in resultados_analisis]
            else:
                df[km["tonoiai"]] = "N/A"
                df[km["tema"]] = "Sector"
                df[km["subtema"]] = "Otras"
            s.update(label="✓ Paso 3 · Análisis Contextual Integrado Completado", state="complete")
            
        rm2 = df.set_index("original_index").to_dict("index")
        for idx, row in enumerate(rows):
            if not row.get("is_duplicate"):
                row.update(rm2.get(row["original_index"], {}))
                
    gc.collect()
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    
    with st.status("Paso 4 · Informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, km)
        st.session_state["output_filename"] = f"Informe_IA_{bn.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name": bn, "brand_aliases": ba,
            "total_rows": len(rows), "unique_rows": len(ta), "duplicates": len(rows) - len(ta),
            "process_duration": f"{time.time() - t0:.0f}s",
            "process_cost": f"${ci + co + ce:.4f} USD",
            "cache_stats": get_embedding_cache().stats()
        })
        s.update(label=f"✓ Completado · {get_embedding_cache().stats()}", state="complete")


# ======================================
# Análisis Rápido (Síncrono)
# ======================================
def run_quick(df, tc, sc, bn, al):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().clear()
    
    # Aplicar la limpieza de menciones en análisis rápido si la columna existe
    if 'Menciones - Empresa' in df.columns:
        df['Menciones - Empresa'] = df['Menciones - Empresa'].apply(limpiar_mencion)
        menciones_col = df['Menciones - Empresa']
    else:
        menciones_col = pd.Series([bn]*len(df))

    df['_txt'] = df.apply(lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))), axis=1)
    
    with st.status("Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
        
    with st.status("Análisis Contextual Integrado", expanded=True) as s:
        pb = st.progress(0)
        res = ClasificadorNoticiasInteligente(bn, al).procesar_lote(
            df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''), menciones_col
        )
        df['Tono IA'] = [r["tono"] for r in res]
        df['Categoría'] = [r["categoria"] for r in res]
        df['Narrativas'] = [r["narrativa"] for r in res]
        s.update(label="✓ Análisis Contextual Integrado", state="complete")
        
    df.drop(columns=['_txt'], inplace=True)
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci + co + ce:.4f} USD"
    return df

def gen_quick_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()

def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Completado</div>'
            '<div class="success-sub">Listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        st.metric("Costo", st.session_state.get('quick_cost', "$0.00"))
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button(
            "Descargar",
            data=gen_quick_excel(st.session_state.quick_result),
            file_name="Analisis_Rapido_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if st.button("Nuevo análisis"):
            for k in ('quick_result', 'quick_df', 'quick_name', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un Excel con columnas de título y resumen.")
        f = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try:
                st.session_state.quick_df   = pd.read_excel(f)
                st.session_state.quick_name = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.success(f"**{st.session_state.quick_name}** cargado")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Col. título",  cols, 0)
            sc = c2.selectbox("Col. resumen", cols, 1 if len(cols) > 1 else 0)
            bn  = st.text_input("Marca",       value="La Cardio")
            bat = st.text_input("Alias (;)",   value="Fundación CardioInfantil;LaCardio;Cardio Infantil;FVDL;Country;Santa Fe;Cardiovascular;Pablo Tobón;Valle de Lily;Shaio")
            if st.form_submit_button("Analizar", use_container_width=True, type="primary"):
                if not bn:
                    st.error("Indica la marca.")
                else:
                    try:
                        openai.api_key = st.secrets["OPENAI_API_KEY"]
                    except:
                        st.error("OPENAI_API_KEY no encontrada.")
                        st.stop()
                    al = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Procesando..."):
                        st.session_state.quick_result = run_quick(st.session_state.quick_df.copy(), tc, sc, bn, al)
                    st.rerun()
        if st.button("Otro archivo"):
            for k in ('quick_df', 'quick_name', 'quick_result', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# Entrada de la Aplicación
# ======================================
def main():
    load_custom_css()
    init_session_state()  # Inicialización segura en tiempo de renderizado de Streamlit
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Noticias - Fundación CardioInfantil</div>
            <div class="app-header-version">v18.2 · Realizado por Johnathan Cortés</div>
        </div>
        <div class="app-header-badge">IA</div>
    </div>""", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["Análisis Completo", "Análisis Rápido"])

    with tab1:
        if not st.session_state.get("processing_complete", False):
            st.markdown('<div class="sec-label">Configuración</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", value="La Cardio", placeholder="Ej: La Cardio", key="bn")
                bat = st.text_input("Alias (separados por ;)", value="Fundación CardioInfantil;LaCardio;Cardio Infantil;FVDL;Country;Santa Fe;Cardiovascular;Pablo Tobón;Valle de Lily;Shaio", placeholder="Ej: FVDL;Country;Santa Fe", key="ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Solo Modelos PKL"],
                    index=0, key="mode"
                )

            tpkl, epkl = None, None
            with st.form("main_form"):
                st.markdown('<div class="sec-label">Archivo de entrada</div>', unsafe_allow_html=True)
                st.markdown("""
                <div class="upload-zone" style="grid-template-columns:1fr">
                    <div class="upload-zone-card">
                        <div class="upload-zone-icon uz-dossier">📋</div>
                        <div class="upload-zone-text">
                            <div class="upload-zone-title">Dossier</div>
                            <div class="upload-zone-desc">Sube las noticias en el formato .xlsx a analizar</div>
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
                f1 = st.file_uploader("Dossier", type=["xlsx"], label_visibility="collapsed", key="f1")

                st.markdown(
                    f'<div class="cluster-info">'
                    f'<b>Consistencia e Inteligencia de Marca</b> · Categorías (Sucesos, Core, Especialidades, Ranking, Sector, Reforma, Corporativo) '
                    f'· Narrativas (Sostenibilidad, Excelencia médica, Innovación + Desarrollo, Marca empleadora, Portafolio, Otras) '
                    f'· Limpieza del prefijo "La Cardio 26 - " y homologación automática de la marca Country en menciones.'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if st.form_submit_button("▶ Iniciar análisis", use_container_width=True, type="primary"):
                    if not all([f1, bn.strip()]):
                        st.error("Por favor completa los campos requeridos e introduce un archivo de Dossier.")
                    else:
                        al = [a.strip() for a in bat.split(";") if a.strip()]
                        cur_mode = st.session_state.get("mode", "API de OpenAI")
                        run_full_process(f1, bn, al, None, None, cur_mode,
                                         xlsx_bytes=None, cliente="", voceros="",
                                         enable_scraping=False)
                        st.rerun()
        else:
            total = st.session_state.total_rows
            uniq  = st.session_state.unique_rows
            dups  = st.session_state.duplicates
            dur   = st.session_state.process_duration
            cost  = st.session_state.get("process_cost", "$0.00")
            st.markdown(
                '<div class="success-banner"><div class="success-icon">✓</div>'
                '<div><div class="success-title">Análisis completado</div>'
                '<div class="success-sub">Informe listo para descargar</div></div></div>',
                unsafe_allow_html=True
            )
            st.markdown(f"""
            <div class="metrics-grid">
              <div class="metric-card m-total"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total</div></div>
              <div class="metric-card m-unique"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicas</div></div>
              <div class="metric-card m-dup"><div class="metric-val" style="color:var(--amber)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
              <div class="metric-card m-time"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo</div></div>
              <div class="metric-card m-cost"><div class="metric-val" style="color:var(--accent)">{cost}</div><div class="metric-lbl">Costo</div></div>
            </div>""", unsafe_allow_html=True)
            if 'cache_stats' in st.session_state: st.caption(f"📊 {st.session_state['cache_stats']}")
            c1, c2 = st.columns(2)
            c1.download_button(
                "⬇ Descargar informe",
                data=st.session_state.output_data,
                file_name=st.session_state.output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            if c2.button("Nuevo análisis", use_container_width=True):
                pwd = st.session_state.get("password_correct")
                st.session_state.clear()
                st.session_state.password_correct = pwd
                st.rerun()

    with tab2:
        render_quick_tab()

    st.markdown(
        '<div class="footer">v18.2 · Análisis de Noticias con IA · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
